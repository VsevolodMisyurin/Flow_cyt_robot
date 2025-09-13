#!/usr/bin/env python
# coding: utf-8

# In[21]:


import os
import re
import pickle
import shutil
import numpy as np
import pandas as pd
import xlwings as xw
from fcsparser import parse
import matplotlib.pyplot as plt
from matplotlib.path import Path
from openpyxl import load_workbook
from matplotlib.patches import Polygon
from catboost import CatBoostClassifier
from openpyxl.utils import get_column_letter
from matplotlib.backends.backend_pdf import PdfPages


# In[22]:


def remove_debris(data):
    """
    Удаляет события, лежащие ниже линии (300000, 300000) — (900000, 0)
    Возвращает отфильтрованный DataFrame
    """
    # Уравнение линии: y = -0.5 * x + 450000
    line_y = lambda x: -0.5 * x + 450000

    # Маска: оставляем точки выше или на линии
    mask = data['SSC-H'] >= line_y(data['FSC-H'])

    # Можно добавить лог, сколько событий удалено
    n_removed = (~mask).sum()
    print(f"Удалено событий как дебрис: {n_removed} из {len(data)}")

    return data[mask].copy()


# In[23]:


def remove_outliers_percentile(df, columns=None, lower_quantile=0.0025, upper_quantile=0.999):
    """
    Удаляет выбросы в датафрейме df, обрезая значения за заданными перцентилями.
    
    Параметры:
    ----------
    df : pd.DataFrame
        Исходный датафрейм.
    columns : list or None
        Список столбцов, по которым ищем выбросы.
        Если None, то берём все числовые столбцы.
    lower_quantile : float
        Нижняя граница (от 0 до 1), по умолчанию 0.0025 (0.25%).
    upper_quantile : float
        Верхняя граница (от 0 до 1), по умолчанию 0.99 (99%).
    
    Возвращает:
    -----------
    pd.DataFrame
        Датафрейм без выбросов.
    """
    
    # Если столбцы не указаны, берём все числовые
    if columns is None:
        columns = df.select_dtypes(include='number').columns.tolist()
    
    # Создадим копию, чтобы не менять исходный датафрейм
    df_clean = df.copy()
    
    for col in columns:
        if col not in df_clean.columns:
            # Если в списке оказался столбец, которого нет в датафрейме
            continue
        
        # Вычислим перцентили
        low_val = df_clean[col].quantile(lower_quantile)
        high_val = df_clean[col].quantile(upper_quantile)
        
        # Фильтруем выбросы
        df_clean = df_clean[(df_clean[col] >= low_val) & (df_clean[col] <= high_val)]
    
    return df_clean


# In[24]:


def log_arcsinh_transform_data(data, cofactor_mapping=None):
    """
    Применяет преобразования к датафрейму.
    
    Для столбцов, в названии которых встречается одно из ключевых слов 
    ('CD'), применяется arcsinh-преобразование с заданным ко-фактором.
    Для остальных столбцов выполняется клиппинг (чтобы избежать log(0)) и логарифмирование.
    
    Параметры:
    ----------
    data : pd.DataFrame
        Исходный датафрейм.
    cofactor_mapping : dict, optional
        Словарь, сопоставляющий ключевые слова и их ко-факторы.
        По умолчанию: {'CD': 5}
    
    Возвращает:
    -----------
    pd.DataFrame
        Датафрейм после преобразований.
    """
    if cofactor_mapping is None:
        cofactor_mapping = {'CD': 3, 'Ig': 3}  # хорошие: 5
    
    data_transformed = data.copy()
    
    for col in data_transformed.columns:
        # Если имя столбца содержит одно из ключевых слов,
        # применяем arcsinh-преобразование с соответствующим ко-фактором.
        applied_arcsinh = False
        for key, cofactor in cofactor_mapping.items():
            if key in col:
                data_transformed[col] = np.arcsinh(data_transformed[col] / cofactor)
                applied_arcsinh = True
                break  # Если нашли соответствие, дальнейшие проверки не требуются.
                
        # Если столбец не содержит заданных ключевых слов, применяем стандартное преобразование:
        if not applied_arcsinh:
            # Клиппинг для предотвращения log(0)
            data_transformed[col] = data_transformed[col].clip(lower=1)
            # Логарифмирование по основанию 2
            data_transformed[col] = np.log2(data_transformed[col])
            
    return data_transformed


# In[32]:


def process_test_data(path, tube, pdf=False, fcs_del=True):
    """
    Обработка тестовых данных внутри Docker-контейнера.

    path: str
        Путь к корневой папке с данными (в контейнере это /app/data).
    tube: list[str]
        Список пробирок (.fcs файлов), в заданном порядке.
    pdf: bool
        Если True, формируется pdf-отчёт.
    fcs_del: bool
        Если True — папка с .fcs будет удалена после обработки.
    """

    # === 1. Папка для отчётов ===
    reports_dir = os.path.join(path, "Reports")
    os.makedirs(reports_dir, exist_ok=True)

    # === 2. Классификаторы и полигоны (всегда в /app/classifiers) ===
    classifiers_root = "/app/classifiers"
    saved_path = classifiers_root  # там и .pkl, и polygons, и media

    # === 3. Собираем список папок для обработки ===
    folders_to_process = []
    for entry in sorted(os.listdir(path)):
        if entry in ["Reports"]:  # в docker нет "Examples" внутри data
            continue
        entry_path = os.path.join(path, entry)
        if not os.path.isdir(entry_path):
            continue

        if entry.isdigit() and len(entry) == 6:
            for sub in sorted(os.listdir(entry_path)):
                sub_path = os.path.join(entry_path, sub)
                if os.path.isdir(sub_path):
                    folders_to_process.append((sub, sub_path))
        else:
            folders_to_process.append((entry, entry_path))

    # === 4. Основной цикл ===
    for folder_name, folder_path in folders_to_process:
        ordered_files = []
        for tube_item in tube:
            tube_prefix = tube_item.split('.')[0]
            files_for_prefix = [
                f for f in os.listdir(folder_path)
                if tube_prefix in f and f.endswith(".fcs")
            ]
            ordered_files.extend(files_for_prefix)

        if not ordered_files:
            continue

        pdf_pages, csv_path, pdf_path = None, None, None
        if pdf:
            pdf_path = os.path.join(reports_dir, f"Report of {folder_name}.pdf")
            pdf_pages = PdfPages(pdf_path)
            csv_path = os.path.join(reports_dir, f"Report_of_{folder_name}.csv")

        print(f"\nОбрабатываю папку: {folder_name}")
        for filename in ordered_files:
            file_path = os.path.join(folder_path, filename)

            # определяем tube_prefix
            tube_prefix = None
            for tube_item in tube:
                prefix = tube_item.split('.')[0]
                if prefix in filename:
                    tube_prefix = prefix
                    break
            if tube_prefix is None:
                print(f"Префикс для {filename} не найден.")
                continue

            # парсинг данных
            try:
                meta, data = parse(file_path)
                del meta
                if len(data) > 120_000:
                    data = data.sample(n=100_000, random_state=42)
                data = data.drop(columns=[col for col in data.columns if col.endswith(('-A', 'Width', 'Time'))])
                data = remove_debris(data)
                data_no_outliers = remove_outliers_percentile(data)
                data_normalized = log_arcsinh_transform_data(data_no_outliers)
            except Exception as e:
                print(f"Ошибка при обработке {filename}: {e}")
                continue

            # загрузка классификатора
            classifier_file = None
            if os.path.isdir(saved_path):
                for clf in os.listdir(saved_path):
                    if clf.endswith(".pkl") and tube_prefix in clf:
                        classifier_file = os.path.join(saved_path, clf)
                        break
            if classifier_file is None:
                print(f"Классификатор для {tube_prefix} не найден в {saved_path}.")
                continue

            try:
                with open(classifier_file, "rb") as f:
                    reducer = pickle.load(f)
                umap_result = reducer.transform(data_normalized)
            except Exception as e:
                print(f"Ошибка при применении UMAP {filename}: {e}")
                continue

            umap_df = pd.DataFrame(
                umap_result,
                columns=[f'UMAP{i+1}' for i in range(reducer.n_components)]
            )

            # полигоны
            polygon_file = os.path.join(saved_path, f"polygons_{tube_prefix}.csv")
            if not os.path.exists(polygon_file):
                print(f"Файл с полигонами {polygon_file} не найден.")
                continue
            try:
                polygons_df = pd.read_csv(polygon_file)
            except Exception as e:
                print(f"Ошибка при загрузке полигонов {polygon_file}: {e}")
                continue

            # визуализация
            fig, ax = plt.subplots(figsize=(10, 8))
            ax.scatter(umap_df["UMAP1"], umap_df["UMAP2"], s=5, alpha=0.5, label="Data")

            umap_df_result = umap_df.copy()
            polygon_groups, added_legend_items = {}, set()

            for (name, n) in polygons_df[["name", "n"]].drop_duplicates().itertuples(index=False):
                poly_data = polygons_df[(polygons_df["name"] == name) & (polygons_df["n"] == n)]
                polygon = np.array(list(zip(poly_data["x"], poly_data["y"])))
                poly_color = poly_data["color"].iloc[0]
                legend_key = (name, poly_color)
                poly_patch = Polygon(polygon, edgecolor=poly_color, facecolor="none", lw=2)
                ax.add_patch(poly_patch)
                if legend_key not in added_legend_items:
                    ax.plot([], [], color=poly_color, label=name)
                    added_legend_items.add(legend_key)
                path_obj = Path(polygon)
                inside = path_obj.contains_points(umap_df[["UMAP1", "UMAP2"]].values).astype(np.int32)
                if name in polygon_groups:
                    polygon_groups[name] |= inside
                else:
                    polygon_groups[name] = inside

            for name, inside_mask in polygon_groups.items():
                umap_df_result[name] = inside_mask

            ax.legend()
            plt.xlabel("UMAP1")
            plt.ylabel("UMAP2")
            plt.title(f"Тестовые данные: {filename}")

            # статистика
            stats = umap_df_result.describe().T.iloc[2:, :2]
            if "Trash" in umap_df_result.columns:
                total_cells = len(umap_df_result)
                trash_count = (umap_df_result["Trash"] == 1).sum()
                granulocytes_count = (umap_df_result["Granulocytes"] == 1).sum() if "Granulocytes" in umap_df_result.columns else 0
                mononuclears = total_cells - trash_count - granulocytes_count
                stats.loc["Mononuclears", ["count", "mean"]] = [mononuclears, mononuclears / total_cells * 100]
                if "Granulocytes" in umap_df_result.columns:
                    stats.loc["Granulocytes", ["count", "mean"]] = [
                        granulocytes_count,
                        granulocytes_count / (total_cells - trash_count) * 100
                    ]
                for pop in polygon_groups.keys():
                    if pop in umap_df_result.columns and pop not in ["Trash", "Granulocytes"]:
                        count = umap_df_result[pop].sum()
                        stats.loc[pop, ["count", "mean"]] = [count, count / mononuclears * 100]

            stats.columns = ['Абс. количество', 'Содержание, %']
            stats['Абс. количество'] = stats['Абс. количество'].astype(np.int32)
            stats['Содержание, %'] = stats['Содержание, %'].astype(np.float32).round(4)
            stats = stats.drop("Trash", errors="ignore")

            print(f"\nСтатистика для пробирки {filename} (пациент: {folder_name}):")
            print(stats)

            # сохранение в PDF и CSV
            if pdf and pdf_pages:
                png_path = os.path.join(reports_dir, f"{folder_name}_{filename}.png")
                fig.savefig(png_path, dpi=100, bbox_inches='tight')
                img = plt.imread(png_path)
                fig_img, ax_img = plt.subplots(figsize=(10, 8))
                ax_img.imshow(img)
                ax_img.axis('off')
                pdf_pages.savefig(fig_img)
                plt.close(fig_img)
                os.remove(png_path)

                fig_table, ax_table = plt.subplots(figsize=(8, 4))
                ax_table.axis('tight')
                ax_table.axis('off')
                table_data = stats.reset_index().values.tolist()
                table_data.insert(0, ["Группа", "Абс. количество", "Содержание, %"])
                ax_table.table(cellText=table_data, loc="center", cellLoc="center")
                plt.title(f"Статистика для образца {folder_name}.")
                pdf_pages.savefig(fig_table)
                plt.close(fig_table)

                stats.to_csv(csv_path, mode='a', header=not os.path.exists(csv_path))
                print(f"Отчёт сохранён в {csv_path}")

        if pdf and pdf_pages:
            pdf_pages.close()
            print(f"Отчёт сохранён в {pdf_path}")

        if fcs_del:
            try:
                shutil.rmtree(folder_path)
                print(f"Удалена папка: {folder_path}")
            except PermissionError as e:
                print(f"Не удалось удалить {folder_path}: {e}")


# In[33]:


# Формирование отчёта, добавление изображений (без xlwings)
def make_report(path, blanks, coordinates, image_folder=None, clean=True):
    """
    Формирует отчёт, перенося данные из CSV в бланк XLSX.
    Перенос данных выполняется с сохранением форматирования ячеек.

    Если передана папка image_folder, то в файл будут вставлены изображения:
      - hl.jpeg — вставляется в верхний левый угол (имитация Header Left)
      - hr.png  — в верхний правый угол (имитация Header Right)
      - fr.jpeg — в нижний правый угол (имитация Footer Right)

    Параметры:
      path         — корневая папка, где находится папка Reports с CSV-отчётами.
      blanks       — папка с шаблонами (бланками) XLSX.
      coordinates  — папка, где лежит файл Coordinates.xlsx с координатами.
      image_folder — (опционально) папка с изображениями.
      clean        — если True, после формирования отчёта CSV-файл удаляется,
                   — если "move", то отчёты переносятся в подпапку Reports/reports_tables
    """

    reports_folder = os.path.join(path, "Reports")
    report_files = [f for f in os.listdir(reports_folder) if f.endswith(".csv")]
    report_names = [f.replace("Report_of_", "").replace(".csv", "") for f in report_files]

    # Загружаем координаты
    coordinates_file = os.path.join(coordinates, "Coordinates.xlsx")
    df_coords = pd.read_excel(coordinates_file, index_col=None)

    # Подготовим пути к картинкам (если нужны)
    if image_folder is not None:
        left_header_img = os.path.join(image_folder, "hl.jpeg")
        right_header_img = os.path.join(image_folder, "hr.png")
        right_footer_img = os.path.join(image_folder, "fr.jpeg")

    for report_name in report_names:
        csv_path = os.path.join(reports_folder, f"Report_of_{report_name}.csv")
        matching_files = [f for f in os.listdir(blanks) if re.match(fr"^{re.escape(report_name)}.*\.xlsx$", f)]
        if not matching_files:
            continue

        xlsx_path = os.path.join(blanks, matching_files[0])
        if not os.path.exists(csv_path):
            continue

        # Загружаем CSV-данные
        df_report = pd.read_csv(csv_path, header=0)
        filter_value = df_report.iloc[0, 0]
        df_filtered = df_coords[df_coords["Col0str1_in_report"] == filter_value]
        if df_filtered.empty:
            continue

        # Загружаем XLSX через openpyxl
        wb = load_workbook(xlsx_path)
        ws = wb.active

        # Сохраняем размеры столбцов/строк
        column_widths = {col: ws.column_dimensions[col].width for col in ws.column_dimensions}
        row_heights = {row: ws.row_dimensions[row].height for row in ws.row_dimensions}

        # Перенос данных
        for _, row in df_filtered.iterrows():
            blank_cell = str(row["Data_in_blank_position"]).strip()
            if pd.isna(row["Data_in_report_position"]):
                ws[blank_cell] = 0
                ws[blank_cell].number_format = '0.00'
            else:
                report_index = int(row["Data_in_report_position"])
                if report_index < len(df_report):
                    value_to_copy = df_report.iloc[report_index, 2]
                    ws[blank_cell] = value_to_copy
                    ws[blank_cell].number_format = '0.00'

        # Восстанавливаем размеры
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        for row, height in row_heights.items():
            ws.row_dimensions[row].height = height

        """
        # Добавляем изображения (имитация header/footer)
        if image_folder is not None:
            from openpyxl.drawing.image import Image as XLImage
            max_col = ws.max_column
            max_row = ws.max_row
            if os.path.exists(left_header_img):
                img = XLImage(left_header_img)
                ws.add_image(img, "A1")
            if os.path.exists(right_header_img):
                img = XLImage(right_header_img)
                ws.add_image(img, f"{get_column_letter(max_col)}1")
            if os.path.exists(right_footer_img):
                img = XLImage(right_footer_img)
                ws.add_image(img, f"{get_column_letter(max_col)}{max_row}")
        """

        wb.save(xlsx_path)
        wb.close()

        # Удаляем/перемещаем CSV
        if clean is True:
            os.remove(csv_path)
        elif clean == "move":
            move_folder = os.path.join(path, "Reports", "reports_tables")
            os.makedirs(move_folder, exist_ok=True)
            shutil.move(csv_path, os.path.join(move_folder, os.path.basename(csv_path)))


# In[34]:


def make_predict(path, blanks, top1_proba_level=0.99, fill_diagnosis=False):
    """
    Достаёт .csv отчёты из Reports, строит датасет, применяет CatBoost классификатор,
    формирует таблицу прогнозов и (опционально) вносит диагноз в XLSX-бланк.

    path   — корневая папка, где Reports
    blanks — папка с шаблонами (куда будут проставляться диагнозы)
    """

    reports_folder = os.path.join(path, "Reports")
    saved_path = os.path.join(path, "Examples", "Saved_UMAP_n_csv")

    report_files = [f for f in os.listdir(reports_folder) if f.endswith(".csv")]
    if not report_files:
        print("CSV отчётов не найдено.")
        return

    results = []

    # Маппинг "фича → модель"
    feature_model_map = {
        "CD19+CD22+CD23-CD5-IgM-": "Chron_full_catboost_model.pkl",
        # можно расширять
    }

    # Словарь диагнозов для автозаполнения
    Chron_dict = {
        'B-ХЛЛ': 'Заключение: иммунофенотип B-клеток соответствует хроническому B-клеточному лимфоидному лейкозу.*',
        'ЛПЗ_нет': 'Заключение: данных за лимфопролиферативное заболевание нет.*',
        'Реакт': 'Заключение: наблюдаемые особенности в наибольшей степени характерны для реактивного процесса.*',
        'ЛМЗ': 'Заключение: иммунофенотип B-клеток соответствует лимфоме из клеток маргинальной зоны.*',
        'МКЛ': 'Заключение: иммунофенотип B-клеток соответствует мантийноклеточной лимфоме.*',
        'ВКЛ': 'Заключение: иммунофенотип B-клеток соответствует волосатоклеточному лейкозу.*'
    }

    for file in report_files:
        csv_path = os.path.join(reports_folder, file)
        try:
            df = pd.read_csv(csv_path)
        except Exception as e:
            print(f"Ошибка чтения {file}: {e}")
            continue

        # Приведение формы таблицы
        try:
            df = df.drop_duplicates(subset=['Unnamed: 0']).T
            df.columns = df.iloc[0]
            df = df.drop(df.index[0])
            df = df.drop(df.index[0])
        except Exception as e:
            print(f"Ошибка при трансформации {file}: {e}")
            continue

        # Конструируем новые признаки
        if all(x in df.columns for x in ["CD4+", "CD8+"]):
            df["imreg"] = df["CD4+"].astype(float) / df["CD8+"].astype(float)

        if all(x in df.columns for x in ["κ+", "λ+"]):
            df["clon"] = np.where(
                (df["κ+"].astype(float) <= 1) & (df["λ+"].astype(float) <= 1),
                1,
                (df["κ+"].astype(float) + 1e-4) / (df["λ+"].astype(float) + 1e-4)
            )

        # Определяем, какой классификатор использовать
        used_model = None
        for feat, model_name in feature_model_map.items():
            if feat in df.columns:
                used_model = model_name
                break

        if not used_model:
            print(f"Для {file} классификатор не найден (нужные фичи отсутствуют).")
            continue

        model_file = os.path.join(saved_path, used_model)
        if not os.path.exists(model_file):
            print(f"Файл модели {model_file} не найден.")
            continue

        # Загружаем CatBoost модель
        try:
            with open(model_file, "rb") as f:
                model = pickle.load(f)
        except Exception as e:
            print(f"Ошибка загрузки модели {model_file}: {e}")
            continue

        # Предсказание
        try:
            preds_proba = model.predict_proba(df)
            classes = model.classes_
        except Exception as e:
            print(f"Ошибка предсказания {file}: {e}")
            continue

        # Берём top-3
        top_idx = np.argsort(-preds_proba[0])[:3]
        patient_name = file.replace("Report_of_", "").replace(".csv", "")
        row = {
            "patient": patient_name,
            "True": None,
            "Top1_class": classes[top_idx[0]],
            "Top1_proba": preds_proba[0][top_idx[0]],
            "Top2_class": classes[top_idx[1]],
            "Top2_proba": preds_proba[0][top_idx[1]],
            "Top3_class": classes[top_idx[2]],
            "Top3_proba": preds_proba[0][top_idx[2]]
        }
        results.append(row)

        # Вписываем диагноз в бланк при уверенности
        if fill_diagnosis and row["Top1_proba"] >= top1_proba_level:
            top1 = row["Top1_class"]
            if top1 in Chron_dict:
                matching_files = [f for f in os.listdir(blanks) if patient_name in f]
                if matching_files:
                    xlsx_path = os.path.join(blanks, matching_files[0])
                    try:
                        wb = load_workbook(xlsx_path)
                        ws = wb.active
                        ws["B48"] = Chron_dict[top1]
                        wb.save(xlsx_path)
                        wb.close()

                        # Переименовываем файл: вставляем диагноз
                        old_name = matching_files[0]
                        parts = old_name.split("_", 2)
                        if len(parts) == 3:
                            new_name = f"{parts[0]}_{top1}_{parts[2]}"
                        else:
                            name, ext = os.path.splitext(old_name)
                            new_name = f"{name}_{top1}{ext}"
                        os.rename(
                            os.path.join(blanks, old_name),
                            os.path.join(blanks, new_name)
                        )
                    except Exception as e:
                        print(f"Ошибка при записи диагноза в {xlsx_path}: {e}")

        # Переносим CSV в Reports/reports_tables
        move_folder = os.path.join(reports_folder, "reports_tables")
        os.makedirs(move_folder, exist_ok=True)
        shutil.move(csv_path, os.path.join(move_folder, os.path.basename(csv_path)))

    # Сохраняем таблицу прогнозов
    if results:
        df_results = pd.DataFrame(results)[[
            "patient", "True",
            "Top1_class", "Top1_proba",
            "Top2_class", "Top2_proba",
            "Top3_class", "Top3_proba"
        ]]
        out_path = os.path.join(blanks, "predictions.xlsx")
        df_results.to_excel(out_path, index=False)
        print(f"Результаты сохранены в {out_path}")
        return df_results
    else:
        print("Прогнозы не построены.")
        return None


# In[ ]:


if __name__ == "__main__":
    # Пути в докере
    data_path = "/app/data"           # где лежит папка Reports и FCS
    blanks_path = "/app/blanks"       # куда сложены бланки заключений
    classifiers_path = "/app/classifiers"  # классификаторы + координаты
    image_folder = f"{classifiers_path}/media"

    # Файлы FCS, которые надо обработать
    tube = [
        "10_56_45_19_138_38.fcs",
        "λ_κ_45_117_138_38.fcs",
        "xx_79a_45_Ki67_3_xx.fcs",
        "23_5_45_22_19_IgM.fcs",
        "103_10_45_34_1a_11c.fcs",
        "xx_56_45_xx_138_38.fcs",
        "λ_κ_45_43_20_xx.fcs",
        "4_8_45_xx_xx_3.fcs"
    ]

    print("=== [1/3] Обработка FCS ===")
    process_test_data(data_path, tube, pdf=True, fcs_del=True)

    print("=== [2/3] Построение отчётов ===")
    make_report(
        path=data_path,
        blanks=blanks_path,
        coordinates=classifiers_path,
        image_folder=image_folder,
        clean=False
    )

    print("=== [3/3] Прогнозы и заполнение диагнозов ===")
    df_results = make_predict(
        path=data_path,
        blanks=blanks_path,
        top1_proba_level=0.94,
        fill_diagnosis=True
    )
    if df_results is not None:
        print(df_results)

    print("=== Анализ завершён успешно ===")

